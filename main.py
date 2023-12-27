import tkinter as tk
from tkinter import filedialog
import pyperclip
import re
import openpyxl


def modified_extract_data(file_path):
    print("Starting data extraction...")
    data = []
    with open(file_path, 'r') as file:
        content = file.read()

        # Find all <valueSettings> blocks
        value_settings_blocks = re.findall(r'<valueSettings>.*?</valueSettings>', content, re.DOTALL)

        for block in value_settings_blocks:
            # Extract all controllingFieldValue instances
            controlling_field_values = re.findall(r'<controllingFieldValue>(.*?)</controllingFieldValue>', block)
            # Extract the single valueName
            value_name_search = re.search(r'<valueName>(.*?)</valueName>', block)
            value_name = value_name_search.group(1) if value_name_search else None

            if controlling_field_values and value_name:
                for controlling_field_value in controlling_field_values:
                    data_row = {
                        'controllingFieldValue': controlling_field_value,
                        'valueName': value_name
                    }
                    print(f"Extracted data row: {data_row}")
                    data.append(data_row)
            else:
                print("A <valueSettings> block without the required subtags was found.")

    print("Data extraction completed.")
    return data


def extract_label_api_data(file_path):
    print("Starting label-api data extraction with field names...")
    data = []
    with open(file_path, 'r') as file:
        content = file.read()

        # Find all <fields> blocks
        fields_blocks = re.findall(r'<fields>.*?</fields>', content, re.DOTALL)

        for block in fields_blocks:
            # Extract the <fullName> of the field
            field_full_name_search = re.search(r'<fullName>(.*?)</fullName>', block)
            field_full_name = field_full_name_search.group(1) if field_full_name_search else None

            # Check if <valueSet> and <valueSetDefinition> exist in the block
            if '<valueSet>' in block and '<valueSetDefinition>' in block:
                # Find all <value> blocks
                value_blocks = re.findall(r'<value>.*?</value>', block, re.DOTALL)

                for value_block in value_blocks:
                    # Extract <fullName> and <label>
                    full_name_search = re.search(r'<fullName>(.*?)</fullName>', value_block)
                    label_search = re.search(r'<label>(.*?)</label>', value_block)

                    full_name = full_name_search.group(1) if full_name_search else None
                    label = label_search.group(1) if label_search else None

                    if full_name and label and field_full_name:
                        data_row = {
                            'fieldFullName': field_full_name,
                            'valueFullName': full_name,
                            'label': label
                        }
                        print(f"Extracted label-api data row: {data_row}")
                        data.append(data_row)
                    else:
                        print("A <value> block without the required fullName and label tags was found.")

    print("Label-api data extraction with field names completed.")
    return data

def extract_recordtype_data(file_path):
    print("Starting recordType data extraction...")
    data = []
    with open(file_path, 'r') as file:
        content = file.read()

        # Find all <recordTypes> blocks
        recordtype_blocks = re.findall(r'<recordTypes>.*?</recordTypes>', content, re.DOTALL)

        for block in recordtype_blocks:
            # Extract <fullName> of the recordType
            recordtype_name_search = re.search(r'<fullName>(.*?)</fullName>', block)
            recordtype_name = recordtype_name_search.group(1) if recordtype_name_search else None

            # Find all <picklistValues> blocks
            picklist_values_blocks = re.findall(r'<picklistValues>.*?</picklistValues>', block, re.DOTALL)

            for picklist_value_block in picklist_values_blocks:
                # Extract <picklist>
                picklist_search = re.search(r'<picklist>(.*?)</picklist>', picklist_value_block)
                picklist = picklist_search.group(1) if picklist_search else None

                # Find all <values> blocks within this <picklistValues> block
                values_blocks = re.findall(r'<values>.*?</values>', picklist_value_block, re.DOTALL)

                for value_block in values_blocks:
                    # Extract <fullName>
                    full_name_search = re.search(r'<fullName>(.*?)</fullName>', value_block)
                    full_name = full_name_search.group(1) if full_name_search else None

                    if recordtype_name and picklist and full_name:
                        data_row = {
                            'recordType': recordtype_name,
                            'field': picklist,
                            'value': full_name
                        }
                        print(f"Extracted recordType data row: {data_row}")
                        data.append(data_row)

    print("RecordType data extraction completed.")
    return data


def extract_fields_data(file_path):
    print("Starting fields data extraction...")
    fields_data = []
    with open(file_path, 'r') as file:
        content = file.read()

        # Find all <fields> blocks
        fields_blocks = re.findall(r'<fields>.*?</fields>', content, re.DOTALL)

        for field_block in fields_blocks:
            # Extract <fullName>, <label>, and <type>
            full_name_search = re.search(r'<fullName>(.*?)</fullName>', field_block)
            label_search = re.search(r'<label>(.*?)</label>', field_block)
            type_search = re.search(r'<type>(.*?)</type>', field_block)

            full_name = full_name_search.group(1) if full_name_search else None
            label = label_search.group(1) if label_search else None
            field_type = type_search.group(1) if type_search else None

            if full_name and label and field_type:
                fields_data_row = {
                    'fullName': full_name,
                    'label': label,
                    'type': field_type
                }
                fields_data.append(fields_data_row)

    print("Fields data extraction completed.")
    return fields_data


def load_file():
    print("Loading file...")
    file_path = file_url_entry.get()
    if not file_path:
        file_path = filedialog.askopenfilename(filetypes=[("Object Files", "*.object")])
        file_url_entry.delete(0, tk.END)
        file_url_entry.insert(0, file_path)

    if file_path:
        print("File path received, processing data...")
        data = modified_extract_data(file_path)
        sorted_data = sorted(data, key=lambda x: x['controllingFieldValue'])

        # Extract fields data which should always run
        fields_data = extract_fields_data(file_path)

        label_api_data = None
        recordtype_data = None

        if labelapi_var.get():
            print("Applying Label-API conversion...")
            label_api_data = extract_label_api_data(file_path)

        if recordtype_var.get():
            print("Extracting RecordType data...")
            recordtype_data = extract_recordtype_data(file_path)

        create_excel_file(sorted_data, label_api_data, recordtype_data, fields_data)
        print(f"Data written to Excel file '{output_file_entry.get()}'.")


def apply_recordtype_dependencies(data):
    print("Applying RecordType dependencies...")
    # Placeholder for RecordType dependencies function
    return data

def apply_label_api_conversion(data):
    print("Applying Label-API conversion...")
    # Placeholder for Label-API conversion function
    return data


def create_excel_file(data, label_api_data=None, recordtype_data=None, fields_data=None):
    output_file_path = output_file_entry.get()
    print(f"Creating Excel file at {output_file_path}...")
    workbook = openpyxl.Workbook()

    # Create and populate the main data sheet
    sheet = workbook.active
    sheet.title = 'Data'
    sheet.append(['controllingFieldValue', 'valueName'])
    for row in data:
        sheet.append([row['controllingFieldValue'], row['valueName']])

    # Create and populate the Fields sheet with field data
    if fields_data:
        fields_sheet = workbook.create_sheet(title="Fields")
        fields_sheet.append(['Field FullName', 'Label', 'Type'])
        for row in fields_data:
            fields_sheet.append([row['fullName'], row['label'], row['type']])

    # Conditionally create and populate the API-Label sheet
    if label_api_data:
        api_label_sheet = workbook.create_sheet(title="API-Label")
        api_label_sheet.append(['Field FullName', 'Value FullName', 'Label'])
        for row in label_api_data:
            api_label_sheet.append([row['fieldFullName'], row['valueFullName'], row['label']])

    # Conditionally create and populate the RecordType sheet
    if recordtype_data:
        recordtype_sheet = workbook.create_sheet(title="RecordType")
        recordtype_sheet.append(['RecordType', 'Field', 'Value'])
        for row in recordtype_data:
            recordtype_sheet.append([row['recordType'], row['field'], row['value']])

    workbook.save(output_file_path)
    print("Excel file created successfully.")


def load_input_file():
    print("Selecting input file...")
    file_path = filedialog.askopenfilename(filetypes=[("Object Files", "*.object")])
    if file_path:
        file_url_entry.delete(0, tk.END)
        file_url_entry.insert(0, file_path)
        print(f"Input file selected: {file_path}")

def load_output_file():
    print("Selecting output file...")
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        output_file_entry.delete(0, tk.END)
        output_file_entry.insert(0, file_path)
        print(f"Output file selected: {file_path}")

# Set up the GUI
print("Setting up the GUI...")
root = tk.Tk()
root.title("Object File Reader")

# Input file selection
file_url_label = tk.Label(root, text="Input File URL:")
file_url_label.pack()
file_url_entry = tk.Entry(root)
file_url_entry.pack()
select_input_file_button = tk.Button(root, text="Select Input File", command=load_input_file)
select_input_file_button.pack()

# Output file selection
output_file_label = tk.Label(root, text="Output Excel File Path:")
output_file_label.pack()
output_file_entry = tk.Entry(root)
output_file_entry.pack()
select_output_file_button = tk.Button(root, text="Select Output File", command=load_output_file)
select_output_file_button.pack()

recordtype_var = tk.IntVar()
recordtype_check = tk.Checkbutton(root, text="RecordType dependencies", variable=recordtype_var)
recordtype_check.pack()

labelapi_var = tk.IntVar()
labelapi_check = tk.Checkbutton(root, text="Label-API conversion", variable=labelapi_var)
labelapi_check.pack()

load_button = tk.Button(root, text="Create Excel File", command=load_file)
load_button.pack()

print("Starting the application...")
root.mainloop()
print("Application closed.")

# TODO: Also extract the Label - API Types. as the DM Tool looks on the label and not the API - ()
# TODO: extract Record Type dependencies ROWS 2388 / 2311 - SIMPLE ()


